from io import BytesIO
from datetime import datetime
from zoneinfo import ZoneInfo
from django.db.models import F, Q
from django.http import HttpResponse
from django.utils import timezone
from django.utils.translation import gettext as _
from rest_framework.viewsets import ReadOnlyModelViewSet
from rest_framework.permissions import IsAuthenticated
from rest_framework.decorators import action
from rest_framework.response import Response
from openpyxl import Workbook
from openpyxl.styles import Font

from abb.utils import get_user_company
from logistic.models import WHStock, WHStockLedger
from logistic.serializers.wms_stock import WHStockSerializer


class WHStockViewSet(ReadOnlyModelViewSet):

    permission_classes = [IsAuthenticated]
    serializer_class = WHStockSerializer
    lookup_field = "uf"

    def get_queryset(self):
        user_company = get_user_company(self.request.user)

        qs = (
            WHStock.objects
            .filter(company=user_company)
            .select_related(
                "product",
                "product__owner",
                "location",
            )
        )

        owner = self.request.query_params.get("owner")
        product = self.request.query_params.get("product")
        location = self.request.query_params.get("location")
        pallet_type = self.request.query_params.get("pallet_type")
        in_stock = self.request.query_params.get("in_stock")


        if owner:
            qs = qs.filter(product__owner__uf=owner)

        if product:
            qs = qs.filter(product__uf=product)

        if location:
            qs = qs.filter(location__uf=location)
        
        print('2114', len(qs), pallet_type)

        if pallet_type:
            qs = qs.filter(pallet_type=pallet_type)

        if in_stock == "true":
            qs = qs.filter(
                Q(quantity__gt=0) |
                Q(pallets__gt=0) |
                Q(area_m2__gt=0) |
                Q(volume_m3__gt=0)
            )

        return qs.order_by(
            "product__name",
            "location__code",
            "pallet_type",
        )
    
    # RECEIVE INBOUND   
    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def movements(self, request):
        company = get_user_company(request.user)

        product = request.GET.get("product")
        owner = request.GET.get("owner")
        location = request.GET.get("location")

        qs = (
            WHStockLedger.objects
            .filter(company=company)
            .select_related("product", "location", "owner")
            .order_by("-created_at")
        )

        if product:
            qs = qs.filter(product__uf=product)

        if owner:
            qs = qs.filter(owner__uf=owner)

        if location:
            qs = qs.filter(location__uf=location)

        data = [
            {
                "id": x.uf,
                "product_name": x.product.name,
                "location_name": x.location.name,
                "owner_name": x.owner.company_name,
                "delta_quantity": x.delta_quantity,
                "delta_pallets": x.delta_pallets,
                "delta_m2": x.delta_area_m2,
                "delta_m3": x.delta_volume_m3,
                "movement_direction": x.movement_direction,
                "source_type": x.source_type,
                "created_at": x.created_at,
            }
            for x in qs
        ]

        return Response(data)

    # EXPORT EXCEL
    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated])
    def export(self, request):
        queryset = self.get_queryset()

        wb = Workbook()
        ws = wb.active
        ws.title = _("stock")

        user_tz = "Europe/Chisinau"
        timezone.activate(ZoneInfo(user_tz))

        now = timezone.localtime(timezone.now())

        # top meta rows
        ws["A1"] = _("stock_export")
        ws["A1"].font = Font(bold=True, size=14)

        ws["A2"] = _("exported_at")
        ws["B2"] = now.strftime("%d-%m-%Y %H:%M")        

        headers = [
            _("owner"),
            _("product_sku"),
            _("product_name"),
            _("location"),
            _("pallet_type"),
            _("quantity"),
            _("pallets"),
            _("area_m2"),
            _("area_m3"),
        ]

        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)

        data_row = header_row + 1

        for row in queryset:
            ws.cell(row=data_row, column=1, value=getattr(row.owner, "company_name", "") if row.owner else "")
            ws.cell(row=data_row, column=2, value=getattr(row.product, "sku", "") if row.product else "")
            ws.cell(row=data_row, column=3, value=getattr(row.product, "name", "") if row.product else "")
            ws.cell(row=data_row, column=4, value=getattr(row.location, "name", "") if row.location else "")
            ws.cell(
                row=data_row,
                column=5,
                value=row.get_pallet_type_display() if row.pallet_type else ""
            )
            ws.cell(row=data_row, column=6, value=float(row.quantity or 0))
            ws.cell(row=data_row, column=7, value=float(row.pallets or 0))
            ws.cell(row=data_row, column=8, value=float(row.area_m2 or 0))
            ws.cell(row=data_row, column=9, value=float(row.volume_m3 or 0))

            data_row += 1

        # simple column widths
        widths = {
            "A": 24,
            "B": 18,
            "C": 30,
            "D": 18,
            "E": 18,
            "F": 12,
            "G": 12,
            "H": 12,
            "I": 12,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename_dt = now.strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="stock_{filename_dt}.xlsx"'
        return response


    # EXPORT STOCK MOVEMENT
    @action(detail=False, methods=["get"], permission_classes=[IsAuthenticated], url_path="movements/export")
    def export_movements(self, request):
        company = get_user_company(request.user)

        qs = (
            WHStockLedger.objects
            .filter(company=company)
            .select_related("product", "location", "owner")
            .order_by("-created_at", "-id")
        )

        owner = request.query_params.get("owner")
        product = request.query_params.get("product")
        location = request.query_params.get("location")

        if owner:
            qs = qs.filter(owner__uf=owner)

        if product:
            qs = qs.filter(product__uf=product)

        if location:
            qs = qs.filter(location__uf=location)

        wb = Workbook()
        ws = wb.active
        ws.title = _("stock_movements")

        user_tz = "Europe/Chisinau"
        timezone.activate(ZoneInfo(user_tz))

        now = timezone.localtime(timezone.now())

        ws["A1"] = _("stock_movements_export")
        ws["A1"].font = Font(bold=True, size=14)

        ws["A2"] = _("exported_at")
        ws["B2"] = now.strftime("%d-%m-%Y %H:%M")

        headers = [
            _("date"),
            _("owner"),
            _("product"),
            _("location"),
            _("direction"),
            _("source"),
            _("pallet_type"),
            _("pallets"),
            _("area_m2"),
            _("area_m3"),
            _("units"),
        ]

        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)

        row_num = header_row + 1

        for row in qs:
            ws.cell(
                row=row_num,
                column=1,
                value=timezone.localtime(row.created_at).strftime("%Y-%m-%d %H:%M:%S")
                if row.created_at else "",
            )
            ws.cell(
                row=row_num,
                column=2,
                value=row.owner.company_name if row.owner else "",
            )
            ws.cell(
                row=row_num,
                column=3,
                value=row.product.name if row.product else "",
            )
            ws.cell(
                row=row_num,
                column=4,
                value=row.location.name if row.location else "",
            )
            ws.cell(
                row=row_num,
                column=5,
                value=_("inbound") if row.movement_direction == "in"
                else _("outbound") if row.movement_direction == "out"
                else "",
            )
            ws.cell(
                row=row_num,
                column=6,
                value=_(row.get_source_type_display()) if row.source_type else "",
            )
            ws.cell(
                row=row_num,
                column=7,
                value=row.get_pallet_type_display() if row.pallet_type else "",
            )
            ws.cell(row=row_num, column=8, value=float(row.delta_pallets or 0))
            ws.cell(row=row_num, column=9, value=float(row.delta_area_m2 or 0))
            ws.cell(row=row_num, column=10, value=float(row.delta_volume_m3 or 0))
            ws.cell(row=row_num, column=11, value=float(row.delta_quantity or 0))

            row_num += 1

        widths = {
            "A": 20,
            "B": 24,
            "C": 30,
            "D": 18,
            "E": 14,
            "F": 16,
            "G": 18,
            "H": 12,
            "I": 12,
            "J": 12,
            "K": 12,
        }
        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename_dt = now.strftime("%Y%m%d_%H%M%S")
        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="stock_movements_{filename_dt}.xlsx"'
        )
        return response