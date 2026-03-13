
from django.db import transaction
from datetime import timedelta
from django.utils import timezone
from django.db.models import Q
from django.shortcuts import get_object_or_404
from django.http import HttpResponse
from django.utils.translation import gettext as _
from django.utils.translation import activate
from django.utils.translation import get_language_from_request
from rest_framework.viewsets import ViewSet
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework.decorators import action
from rest_framework import status
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from abb.utils import get_user_company
from att.models import Contact
from logistic.models import WHContactTariffHandlingTierOverride, WHStorageBillingMode, WHTariff, WHContactTariffOverride
from logistic.serializers.wms_tariff import WHTariffSerializer


class WHTariffViewSet(ViewSet):

    permission_classes = [IsAuthenticated]

    def _parse_date(self, value):
        if not value:
            return None
        return timezone.datetime.strptime(value, "%Y-%m-%d").date()

    def _get_active_override(self, company, contact, as_of_date=None):
        if as_of_date is None:
            as_of_date = timezone.localdate()

        return (
            WHContactTariffOverride.objects
            .filter(
                company=company,
                contact=contact,
                period_start__lte=as_of_date,
            )
            .filter(
                Q(period_end__isnull=True) | Q(period_end__gte=as_of_date) 
            )
            .prefetch_related("handling_tier_overrides")
            .order_by("-period_start", "-created_at")
            .first()
        )

    def _get_default_tariff(self, company):
        default_tariff = (
            WHTariff.objects
            .filter(company=company, is_active=True)
            .first()
        )

        if default_tariff is None:
            default_tariff = WHTariff.objects.create(
                company=company,
                storage_mode=WHStorageBillingMode.PALLET,
                storage_per_unit_per_day=0,
                storage_per_m2_per_day=0,
                storage_per_m3_per_day=0,
                storage_per_euro_pallet_per_day=0,
                storage_per_iso2_pallet_per_day=0,
                storage_per_block_pallet_per_day=0,
                # inbound_per_line=0,
                # outbound_per_order=0,
                # outbound_per_line=0,
                storage_min_days=1,
                handling_tier_mode="bracket",
                is_active=True,
            )

        return default_tariff

    def _extract_override_defaults(self, data):
        def empty_to_none(value):
            return None if value == "" else value

        return {
            "storage_mode": empty_to_none(data.get("storage_mode")),
            "storage_per_unit_per_day": empty_to_none(data.get("storage_per_unit_per_day")),
            "storage_per_m2_per_day": empty_to_none(data.get("storage_per_m2_per_day")),
            "storage_per_m3_per_day": empty_to_none(data.get("storage_per_m3_per_day")),
            "storage_per_euro_pallet_per_day": empty_to_none(data.get("storage_per_euro_pallet_per_day")),
            "storage_per_iso2_pallet_per_day": empty_to_none(data.get("storage_per_iso2_pallet_per_day")),
            "storage_per_block_pallet_per_day": empty_to_none(data.get("storage_per_block_pallet_per_day")),
            "storage_min_days": empty_to_none(data.get("storage_min_days")),
            # "inbound_per_line": empty_to_none(data.get("inbound_per_line")),
            # "outbound_per_order": empty_to_none(data.get("outbound_per_order")),
            # "outbound_per_line": empty_to_none(data.get("outbound_per_line")),
            "handling_tier_mode": empty_to_none(data.get("handling_tier_mode")),
        }
    
    def _sync_handling_tier_overrides(self, override, handling_tiers):
        """
        Replace override handling tiers with submitted payload.
        Ignore incomplete rows.
        """
        WHContactTariffHandlingTierOverride.objects.filter(override=override).delete()

        def empty_to_none(value):
            return None if value in ("", None) else value

        rows_to_create = []

        for index, row in enumerate(handling_tiers or []):
            fee_type = empty_to_none(row.get("fee_type"))
            unit = empty_to_none(row.get("unit"))
            min_quantity = empty_to_none(row.get("min_quantity"))
            max_quantity = empty_to_none(row.get("max_quantity"))
            price = empty_to_none(row.get("price"))
            order = row.get("order", index)

            # skip incomplete rows
            if not fee_type or not unit or min_quantity is None or price is None:
                continue

            rows_to_create.append(
                WHContactTariffHandlingTierOverride(
                    override=override,
                    fee_type=fee_type,
                    unit=unit,
                    min_quantity=min_quantity,
                    max_quantity=max_quantity,
                    price=price,
                    order=order,
                )
            )

        if rows_to_create:
            WHContactTariffHandlingTierOverride.objects.bulk_create(rows_to_create)

    def _get_previous_override(self, company, contact, period_start):
        return (
            WHContactTariffOverride.objects
            .select_for_update()
            .filter(company=company, contact=contact, period_start__lt=period_start)
            .order_by("-period_start", "-created_at")
            .first()
        )

    def _get_next_override(self, company, contact, period_start):
        return (
            WHContactTariffOverride.objects
            .select_for_update()
            .filter(company=company, contact=contact, period_start__gt=period_start)
            .order_by("period_start", "created_at")
            .first()
        )

    def list(self, request):
        company = get_user_company(request.user)
        today = timezone.localdate()
        default_tariff = self._get_default_tariff(company)

        overrides = (
            WHContactTariffOverride.objects
            .filter(company=company, period_start__lte=today)
            .filter(Q(period_end__isnull=True) | Q(period_end__gte=today))
            .select_related("contact")
            .prefetch_related("handling_tier_overrides")
            .order_by("contact_id", "-period_start", "-created_at")
        )

        active_override_by_contact_id = {}
        for override in overrides:
            if override.contact_id not in active_override_by_contact_id:
                active_override_by_contact_id[override.contact_id] = override

        data = []

        for override in active_override_by_contact_id.values():
            contact = override.contact

            storage_mode = override.storage_mode or default_tariff.storage_mode

            storage_per_euro_pallet = (
                override.storage_per_euro_pallet_per_day
                if override.storage_per_euro_pallet_per_day is not None
                else default_tariff.storage_per_euro_pallet_per_day
            )

            storage_per_iso2_pallet = (
                override.storage_per_iso2_pallet_per_day
                if override.storage_per_iso2_pallet_per_day is not None
                else default_tariff.storage_per_iso2_pallet_per_day
            )

            storage_per_block_pallet = (
                override.storage_per_block_pallet_per_day
                if override.storage_per_block_pallet_per_day is not None
                else default_tariff.storage_per_block_pallet_per_day
            )

            storage_per_unit = (
                override.storage_per_unit_per_day
                if override.storage_per_unit_per_day is not None
                else default_tariff.storage_per_unit_per_day
            )

            storage_per_m2 = (
                override.storage_per_m2_per_day
                if override.storage_per_m2_per_day is not None
                else default_tariff.storage_per_m2_per_day
            )

            storage_per_m3 = (
                override.storage_per_m3_per_day
                if override.storage_per_m3_per_day is not None
                else default_tariff.storage_per_m3_per_day
            )

            storage_min_days = (
                override.storage_min_days
                if override.storage_min_days is not None
                else default_tariff.storage_min_days
            )

            # inbound = (
            #     override.inbound_per_line
            #     if override.inbound_per_line is not None
            #     else default_tariff.inbound_per_line
            # )

            # outbound_order = (
            #     override.outbound_per_order
            #     if override.outbound_per_order is not None
            #     else default_tariff.outbound_per_order
            # )

            # outbound_line = (
            #     override.outbound_per_line
            #     if override.outbound_per_line is not None
            #     else default_tariff.outbound_per_line
            # )

            handling_tier_mode = (
                override.handling_tier_mode
                if override.handling_tier_mode
                else default_tariff.handling_tier_mode
            )

            handling_tiers = [
                {
                    "fee_type": tier.fee_type,
                    "unit": tier.unit,
                    "min_quantity": tier.min_quantity,
                    "max_quantity": tier.max_quantity,
                    "price": tier.price,
                    "order": tier.order,
                }
                for tier in override.handling_tier_overrides.all()
            ]

            data.append({
                "contact": contact.uf,
                "contact_name": contact.company_name,
                "period_start": override.period_start,
                "period_end": override.period_end,
                "is_active": override.is_active,
                "storage_mode": storage_mode,
                "storage_per_euro_pallet_per_day": storage_per_euro_pallet,
                "storage_per_iso2_pallet_per_day": storage_per_iso2_pallet,
                "storage_per_block_pallet_per_day": storage_per_block_pallet,
                "storage_per_unit_per_day": storage_per_unit,
                "storage_per_m2_per_day": storage_per_m2,
                "storage_per_m3_per_day": storage_per_m3,
                "storage_min_days": storage_min_days,
                # "inbound_per_line": inbound,
                # "outbound_per_order": outbound_order,
                # "outbound_per_line": outbound_line,
                "handling_tier_mode": handling_tier_mode,
                "handling_tiers": handling_tiers,
                "is_override": True,
            })

        serializer = WHTariffSerializer(data, many=True)
        return Response(serializer.data)
    
    @action(detail=False, methods=["post"])
    @transaction.atomic
    def create_override(self, request):
        company = get_user_company(request.user)
        data = request.data.copy()

        contact_uf = data.get("contact")
        contact = get_object_or_404(Contact, company=company, uf=contact_uf)

        period_start = self._parse_date(data.get("period_start")) or timezone.localdate()
        period_end = self._parse_date(data.get("period_end"))

        defaults = self._extract_override_defaults(data)

        previous_override = self._get_previous_override(company, contact, period_start)
        next_override = self._get_next_override(company, contact, period_start)

        if previous_override:
            previous_override.period_end = period_start - timedelta(days=1)
            previous_override.save(update_fields=["period_end"])

        if period_end is None and next_override:
            period_end = next_override.period_start - timedelta(days=1)

        override = WHContactTariffOverride.objects.create(
            company=company,
            contact=contact,
            period_start=period_start,
            period_end=period_end,
            **defaults,
        )

        self._sync_handling_tier_overrides(
            override=override,
            handling_tiers=data.get("handling_tiers") or [],
        )

        return Response({"success": True}, status=status.HTTP_200_OK) 
    
    @action(detail=True, methods=["patch"])
    @transaction.atomic
    def update_override(self, request, pk=None):
        company = get_user_company(request.user)
        contact = get_object_or_404(Contact, company=company, uf=pk)

        new_start_date = self._parse_date(request.data.get("period_start")) or timezone.localdate()
        new_period_end = self._parse_date(request.data.get("period_end"))

        defaults = self._extract_override_defaults(request.data)

        existing_same_start = (
            WHContactTariffOverride.objects
            .select_for_update()
            .filter(
                company=company,
                contact=contact,
                period_start=new_start_date,
            )
            .first()
        )

        if existing_same_start:
            for field, value in defaults.items():
                if field in request.data:
                    setattr(existing_same_start, field, value)

            if "period_end" in request.data:
                existing_same_start.period_end = new_period_end

            existing_same_start.save()

            if "handling_tiers" in request.data:
                self._sync_handling_tier_overrides(
                    override=existing_same_start,
                    handling_tiers=request.data.get("handling_tiers") or [],
                )

            return Response({"status": "ok"}, status=status.HTTP_200_OK)

        previous_override = self._get_previous_override(company, contact, new_start_date)
        next_override = self._get_next_override(company, contact, new_start_date)

        if previous_override:
            previous_override.period_end = new_start_date - timedelta(days=1)
            previous_override.save(update_fields=["period_end"])

        if new_period_end is None and next_override:
            new_period_end = next_override.period_start - timedelta(days=1)

        override = WHContactTariffOverride.objects.create(
            company=company,
            contact=contact,
            period_start=new_start_date,
            period_end=new_period_end,
            **defaults,
        )

        self._sync_handling_tier_overrides(
            override=override,
            handling_tiers=request.data.get("handling_tiers") or [],
        )

        return Response({"status": "ok"}, status=status.HTTP_200_OK)
    
    @action(detail=True, methods=["get"])
    def override_history(self, request, pk=None):
        company = get_user_company(request.user)
        contact = get_object_or_404(Contact, company=company, uf=pk)
        today = timezone.localdate()
        default_tariff = self._get_default_tariff(company)

        overrides = (
            WHContactTariffOverride.objects
            .filter(company=company, contact=contact)
            .prefetch_related("handling_tier_overrides")
            .order_by("-period_start", "-created_at")
        )

        data = []
        for override in overrides:
            is_active = (
                override.period_start <= today and
                (override.period_end is None or override.period_end >= today)
            )

            storage_mode = override.storage_mode or default_tariff.storage_mode

            storage_per_euro_pallet = (
                override.storage_per_euro_pallet_per_day
                if override.storage_per_euro_pallet_per_day is not None
                else default_tariff.storage_per_euro_pallet_per_day
            )

            storage_per_iso2_pallet = (
                override.storage_per_iso2_pallet_per_day
                if override.storage_per_iso2_pallet_per_day is not None
                else default_tariff.storage_per_iso2_pallet_per_day
            )

            storage_per_block_pallet = (
                override.storage_per_block_pallet_per_day
                if override.storage_per_block_pallet_per_day is not None
                else default_tariff.storage_per_block_pallet_per_day
            )

            storage_per_unit = (
                override.storage_per_unit_per_day
                if override.storage_per_unit_per_day is not None
                else default_tariff.storage_per_unit_per_day
            )

            storage_per_m2 = (
                override.storage_per_m2_per_day
                if override.storage_per_m2_per_day is not None
                else default_tariff.storage_per_m2_per_day
            )

            storage_per_m3 = (
                override.storage_per_m3_per_day
                if override.storage_per_m3_per_day is not None
                else default_tariff.storage_per_m3_per_day
            )

            storage_min_days = (
                override.storage_min_days
                if override.storage_min_days is not None
                else default_tariff.storage_min_days
            )

            inbound_per_line = (
                override.inbound_per_line
                if override.inbound_per_line is not None
                else default_tariff.inbound_per_line
            )

            outbound_per_order = (
                override.outbound_per_order
                if override.outbound_per_order is not None
                else default_tariff.outbound_per_order
            )

            outbound_per_line = (
                override.outbound_per_line
                if override.outbound_per_line is not None
                else default_tariff.outbound_per_line
            )

            handling_tier_mode = (
                override.handling_tier_mode
                if override.handling_tier_mode
                else default_tariff.handling_tier_mode
            )

            data.append({
                "uf": override.uf,
                "contact": contact.uf,
                "contact_name": contact.company_name,
                "period_start": override.period_start,
                "period_end": override.period_end,
                "is_active": is_active,
                "storage_mode": storage_mode,
                "storage_per_euro_pallet_per_day": storage_per_euro_pallet,
                "storage_per_iso2_pallet_per_day": storage_per_iso2_pallet,
                "storage_per_block_pallet_per_day": storage_per_block_pallet,
                "storage_per_unit_per_day": storage_per_unit,
                "storage_per_m2_per_day": storage_per_m2,
                "storage_per_m3_per_day": storage_per_m3,
                "storage_min_days": storage_min_days,
                "inbound_per_line": inbound_per_line,
                "outbound_per_order": outbound_per_order,
                "outbound_per_line": outbound_per_line,
                "handling_tier_mode": handling_tier_mode,
                "handling_tiers": [
                    {
                        "fee_type": tier.fee_type,
                        "unit": tier.unit,
                        "min_quantity": tier.min_quantity,
                        "max_quantity": tier.max_quantity,
                        "price": tier.price,
                        "order": tier.order,
                    }
                    for tier in override.handling_tier_overrides.all()
                ],
                "is_override": True,
            })

        serializer = WHTariffSerializer(data, many=True)
        return Response(serializer.data)

    ### EXPORT ###
    @action(detail=True, methods=["get"], permission_classes=[IsAuthenticated], url_path="override-history/export")
    def export_override_history(self, request, pk=None):
        lang = get_language_from_request(request)
        activate(lang)

        print('LANG5044', lang)

        company = get_user_company(request.user)
        contact = get_object_or_404(Contact, company=company, uf=pk)

        overrides = (
            WHContactTariffOverride.objects
            .filter(company=company, contact=contact)
            .prefetch_related("handling_tier_overrides")
            .order_by("-period_start", "-created_at")
        )

        wb = Workbook()
        ws = wb.active
        ws.title = _("tariff_history")

        headers = [
            _("contact"),
            _("period_start"),
            _("period_end"),
            _("is_active"),
            _("storage_mode"),
            _("storage_min_days"),
            _("euro_pallet_per_day"),
            _("iso2_pallet_per_day"),
            _("block_pallet_per_day"),
            _("m2_per_day"),
            _("m3_per_day"),
            _("unit_per_day"),            
            _("handling_tier_mode"),
           _("handling_tiers_count"),
            _("handling_tiers"),
        ]

        ws.append(headers)

        header_fill = PatternFill("solid", fgColor="D9EAF7")
        header_font = Font(bold=True)

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(vertical="center", horizontal="center")

        today = timezone.localdate()

        for override in overrides:
            is_active = (
                override.period_start <= today and
                (override.period_end is None or override.period_end >= today)
            )

            tiers_text = "\n".join([
                f"{tier.fee_type} | {tier.unit} | min={tier.min_quantity} | max={tier.max_quantity or '-'} | price={tier.price}"
                for tier in override.handling_tier_overrides.all()
            ])

            ws.append([
                contact.company_name,
                override.period_start.strftime("%d-%m-%Y") if override.period_start else "",
                override.period_end.strftime("%d-%m-%Y") if override.period_end else "",
                "Yes" if is_active else "No",
                override.storage_mode or "",
                override.storage_min_days if override.storage_min_days is not None else "",
                override.storage_per_euro_pallet_per_day if override.storage_per_euro_pallet_per_day is not None else "",
                override.storage_per_iso2_pallet_per_day if override.storage_per_iso2_pallet_per_day is not None else "",
                override.storage_per_block_pallet_per_day if override.storage_per_block_pallet_per_day is not None else "",
                override.storage_per_m2_per_day if override.storage_per_m2_per_day is not None else "",
                override.storage_per_m3_per_day if override.storage_per_m3_per_day is not None else "",
                override.storage_per_unit_per_day if override.storage_per_unit_per_day is not None else "",
               
                override.handling_tier_mode or "",
                override.handling_tier_overrides.count(),
                tiers_text,
            ])

        widths = {
            "A": 24,
            "B": 14,
            "C": 14,
            "D": 10,
            "E": 16,
            "F": 16,
            "G": 18,
            "H": 18,
            "I": 18,
            "J": 12,
            "K": 12,
            "L": 12,
            "M": 22,
            "N": 24,
            "O": 60,

        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename = f"tariff_history_{contact.company_name}_{contact.uf}.xlsx"

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = f'attachment; filename="{filename}"'
        return response
