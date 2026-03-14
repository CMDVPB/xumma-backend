from decimal import Decimal
from io import BytesIO
from zoneinfo import ZoneInfo
from datetime import datetime
from django.utils import timezone
from django.db.models import F, Q
from django.shortcuts import get_object_or_404
from django.db.models import Prefetch, Sum
from django.db import transaction
from rest_framework.permissions import IsAuthenticated
from rest_framework.response import Response
from rest_framework import status
from rest_framework.decorators import action
from rest_framework import viewsets
from django.utils.translation import gettext as _, override
from rest_framework.decorators import action
from django.http import HttpResponse
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side
from openpyxl.worksheet.page import PageMargins
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.properties import PageSetupProperties

from abb.utils import get_user_company
from att.models import Contact
from logistic.models import WHBillingCharge, WHBillingInvoice, WHBillingInvoiceLine, WHBillingPeriod
from logistic.serializers.wms_billing import WHBillingInvoiceManualLinesCreateSerializer, WHBillingInvoiceSerializer, WHBillingPeriodSerializer, WHIssueBillingDocumentsSerializer
from logistic.services.wms_billing_documents import issue_billing_documents_for_period
from logistic.services.wms_billing_engine import (generate_storage_billing_for_period, regenerate_all_billing_charges_for_period, 
                                                    )


###### HELPER ######
def _get_or_create_current_period(company):

    today = timezone.localdate()

    last_period = (
        WHBillingPeriod.objects
        .filter(company=company)
        .order_by("-end_date")
        .first()
    )

    if not last_period:
        return WHBillingPeriod.objects.create(
            company=company,
            start_date=today.replace(day=1),
            end_date=today
        )

    # if period already invoiced → create new
    has_invoice = WHBillingInvoice.objects.filter(
        company=company,
        period=last_period
    ).exists()

    if has_invoice:
        return WHBillingPeriod.objects.create(
            company=company,
            start_date=last_period.end_date,
            end_date=today
        )

    if last_period.end_date == today:
        return last_period

    last_period.end_date = today
    last_period.save(update_fields=["end_date"])

    return last_period


def _translated_charge_type_label(charge_type):
    return {
        "storage": _("storage"),
        "inbound": _("inbouned"),
        "outbound_order": _("outbound_order"),
        "outbound_line": _("outbound_line"),
        "handling_loading": _("handling_loading"),
        "handling_unloading": _("handling_unloading"),
    }.get(charge_type, charge_type or "")


def _translated_line_description(line):
    charge_label = _translated_charge_type_label(line.charge_type)

    charge = line.charges.first() if hasattr(line, "charges") else None

    parts = [charge_label]

    if charge and charge.product:
        parts.append(f'{_("product")}: {charge.product.name}')

    if charge and charge.location:
        location_label = charge.location.name or charge.location.code or ""
        parts.append(f'{_("location")}: {location_label}')

    if charge and charge.billing_period:
        parts.append(
            f'{_("period")}: {charge.billing_period.start_date.strftime("%d-%m-%Y")} '
            f'→ {charge.billing_period.end_date.strftime("%d-%m-%Y")}'
        )

    return " | ".join(parts)


class WHBillingPeriodViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]    
    serializer_class = WHBillingPeriodSerializer
    lookup_field = "uf"

    def get_queryset(self):
        company = get_user_company(self.request.user)
        return (
            WHBillingPeriod.objects
            .filter(company=company)
            .order_by("-start_date", "-created_at")
        )

    @action(detail=True, methods=["post"], url_path="generate-charges")
    def generate_charges(self, request, uf=None):
        company = get_user_company(request.user)

        period = get_object_or_404(
            WHBillingPeriod,
            uf=uf,
            company=company,
        )

        if period.is_closed:
            return Response(
                {"detail": "Billing period is closed"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        try:
            result = regenerate_all_billing_charges_for_period(
                company=company,
                period=period,
            )
        except ValueError as exc:
            return Response(
                {"detail": str(exc)},
                status=status.HTTP_400_BAD_REQUEST,
            )
        
        return Response(
            {
                "detail": "Charges generated successfully",
                "created_count": result["created_count"],
                "storage_count": len(result["storage"]),
                "handling_count": len(result["handling"]),
            },
            status=status.HTTP_200_OK,
        )

    @action(detail=True, methods=["post"], url_path="close")
    def close_period(self, request, uf=None):
        period = self.get_object()

        if period.is_closed:
            return Response(
                {"detail": "Period already closed"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        period.is_closed = True
        period.save(update_fields=["is_closed"])

        return Response({"detail": "Billing period closed"})
    

class WHBillingDocumentViewSet(viewsets.ModelViewSet):
    permission_classes = [IsAuthenticated]
    serializer_class = WHBillingInvoiceSerializer
    lookup_field = "uf"

    def get_queryset(self):
        company = get_user_company(self.request.user)

        queryset = (
            WHBillingInvoice.objects
            .filter(company=company)
            .select_related("contact", "period")
            .prefetch_related("invoice_wh_lines")
            .order_by("-created_at")
        )

        period = self.request.query_params.get("period")
        contact = self.request.query_params.get("contact")
        status_param = self.request.query_params.get("status")

        if period:
            queryset = queryset.filter(period__uf=period)

        if contact:
            queryset = queryset.filter(contact__uf=contact)

        if status_param:
            queryset = queryset.filter(status=status_param)

        return queryset

    @action(detail=False, methods=["post"], url_path="issue-for-period")
    @transaction.atomic
    def issue_for_period(self, request):
        company = get_user_company(request.user)

        serializer = WHIssueBillingDocumentsSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        period_uf = serializer.validated_data["period"]
        contact_ufs = serializer.validated_data.get("contacts") or []

        period = get_object_or_404(
            WHBillingPeriod,
            uf=period_uf,
            company=company,
        )

        if period.is_closed:
            return Response(
                {"detail": "Billing period is closed"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        contact_ids = None
        if contact_ufs:
            contact_ids = list(
                Contact.objects.filter(uf__in=contact_ufs).values_list("id", flat=True)
            )

        documents = issue_billing_documents_for_period(
            company=company,
            period=period,
            contact_ids=contact_ids,
        )

        return Response(
            {
                "detail": "Billing documents created",
                "count": len(documents),
                "documents": [doc.uf for doc in documents],
            },
            status=status.HTTP_201_CREATED,
        )

    @action(detail=True, methods=["post"])
    def issue(self, request, uf=None):
        document = self.get_object()

        if document.status != "draft":
            return Response(
                {"detail": "Only draft documents can be issued"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        document.status = "issued"
        document.save(update_fields=["status"])

        return Response({"status": "issued"})

    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        document = self.get_object()

        if document.status != "draft":
            return Response(
                {"detail": "Only draft documents can be deleted"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        charge_ids = list(
            WHBillingCharge.objects.filter(
                invoice_lines__invoice=document
            ).values_list("id", flat=True)
        )

        if charge_ids:
            WHBillingCharge.objects.filter(id__in=charge_ids).update(invoiced=False)

        document.delete()
        return Response(status=status.HTTP_204_NO_CONTENT)

    # EXPORT EXCEL LIST
    @action(detail=False, methods=["get"], url_path="export")
    def export_documents(self, request):
        company = get_user_company(request.user)

        qs = (
            WHBillingInvoice.objects
            .filter(company=company)
            .select_related("contact", "period")
            .prefetch_related("invoice_wh_lines")
            .order_by("-created_at", "-id")
        )

        period = request.query_params.get("period")
        contact = request.query_params.get("contact")
        status_param = request.query_params.get("status")

        if period:
            qs = qs.filter(period__uf=period)

        if contact:
            qs = qs.filter(contact__uf=contact)

        if status_param:
            qs = qs.filter(status=status_param)

        wb = Workbook()
        ws = wb.active
        ws.title = _("billing_documents")

        user_tz = "Europe/Chisinau"
        timezone.activate(ZoneInfo(user_tz))
        now = timezone.localtime(timezone.now())

        ws["A1"] = _("billing_documents_export")
        ws["A1"].font = Font(bold=True, size=14)

        ws["A2"] = _("exported_at")
        ws["B2"] = now.strftime("%d-%m-%Y %H:%M")

        headers = [
            _("document_no"),
            _("customer"),
            _("period_start"),
            _("period_end"),
            _("status"),
            _("created_at"),
            _("total_amount"),
            _("line_type"),
            _("line_description"),
            _("quantity"),
            _("unit_price"),
            _("line_total"),
        ]

        header_row = 4
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)

        status_map = {
            "draft": _("draft"),
            "issued": _("issued"),
            "paid": _("paid"),
        }

        charge_type_map = {
            "storage": _("storage"),
            "inbound": _("inbound"),
            "outbound_order": _("outbound_order"),
            "outbound_line": _("outbound_line"),
            "handling_loading": _("handling_loading"),
            "handling_unloading": _("handling_unloading"),
        }

        row_num = header_row + 1

        for document in qs:
            lines = list(document.invoice_wh_lines.all())

            if not lines:
                ws.cell(row=row_num, column=1, value=document.uf)
                ws.cell(
                    row=row_num,
                    column=2,
                    value=document.contact.company_name if document.contact else "",
                )
                ws.cell(
                    row=row_num,
                    column=3,
                    value=document.period.start_date.strftime("%d-%m-%Y") if document.period else "",
                )
                ws.cell(
                    row=row_num,
                    column=4,
                    value=document.period.end_date.strftime("%d-%m-%Y") if document.period else "",
                )
                ws.cell(
                    row=row_num,
                    column=5,
                    value=status_map.get(document.status, document.status or ""),
                )
                ws.cell(
                    row=row_num,
                    column=6,
                    value=timezone.localtime(document.created_at).strftime("%Y-%m-%d %H:%M:%S")
                    if document.created_at else "",
                )
                ws.cell(row=row_num, column=7, value=float(document.total_amount or 0))
                row_num += 1
                continue

            for line in lines:
                ws.cell(row=row_num, column=1, value=document.uf)
                ws.cell(
                    row=row_num,
                    column=2,
                    value=document.contact.company_name if document.contact else "",
                )
                ws.cell(
                    row=row_num,
                    column=3,
                    value=document.period.start_date.strftime("%d-%m-%Y") if document.period else "",
                )
                ws.cell(
                    row=row_num,
                    column=4,
                    value=document.period.end_date.strftime("%d-%m-%Y") if document.period else "",
                )
                ws.cell(
                    row=row_num,
                    column=5,
                    value=status_map.get(document.status, document.status or ""),
                )
                ws.cell(
                    row=row_num,
                    column=6,
                    value=timezone.localtime(document.created_at).strftime("%Y-%m-%d %H:%M:%S")
                    if document.created_at else "",
                )
                ws.cell(row=row_num, column=7, value=float(document.total_amount or 0))
                ws.cell(
                    row=row_num,
                    column=8,
                    value=charge_type_map.get(line.charge_type, line.charge_type or ""),
                )
                ws.cell(row=row_num, column=9, value=line.description or "")
                ws.cell(row=row_num, column=10, value=float(line.quantity or 0))
                ws.cell(row=row_num, column=11, value=float(line.unit_price or 0))
                ws.cell(row=row_num, column=12, value=float(line.total or 0))

                row_num += 1

        widths = {
            "A": 38,
            "B": 28,
            "C": 14,
            "D": 14,
            "E": 14,
            "F": 20,
            "G": 14,
            "H": 20,
            "I": 40,
            "J": 12,
            "K": 12,
            "L": 14,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename_dt = now.strftime("%Y%m%d_%H%M%S")

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="billing_documents_{filename_dt}.xlsx"'
        )
        return response
    
    # EXPORT EXCEL BILL DETAILS
    @action(detail=True, methods=["get"], url_path="export")
    def export_document(self, request, uf=None):
        document = self.get_object()

        wb = Workbook()
        ws = wb.active
        ws.title = _("billing_document")

        user_tz = "Europe/Chisinau"
        timezone.activate(ZoneInfo(user_tz))
        now = timezone.localtime(timezone.now())

        ws["A1"] = _("billing_document_export")
        ws["A1"].font = Font(bold=True, size=14)

        ws["A2"] = _("exported_at")
        ws["B2"] = now.strftime("%d-%m-%Y %H:%M")

        ws["A3"] = _("document_no")
        ws["B3"] = document.uf

        ws["A4"] = _("customer")
        ws["B4"] = document.contact.company_name if document.contact else ""

        ws["A5"] = _("period_start")
        ws["B5"] = document.period.start_date.strftime("%d-%m-%Y") if document.period else ""

        ws["A6"] = _("period_end")
        ws["B6"] = document.period.end_date.strftime("%d-%m-%Y") if document.period else ""

        status_map = {
            "draft": _("draft"),
            "issued": _("issued"),
            "paid": _("paid"),
        }

        ws["A7"] = _("status")
        ws["B7"] = status_map.get(document.status, document.status or "")

        ws["A8"] = _("created_at")
        ws["B8"] = (
            timezone.localtime(document.created_at).strftime("%d-%m-%Y %H:%M")
            if document.created_at else ""
        )

        headers = [
            _("line_type"),
            _("line_description"),
            _("quantity"),
            _("unit_price"),
            _("line_total"),
        ]

        header_row = 11
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)

        charge_type_map = {
            "storage": _("storage"),
            "inbound": _("inbound"),
            "outbound_order": _("outbound_order"),
            "outbound_line": _("outbound_line"),
            "handling_loading": _("handling_loading"),
            "handling_unloading": _("handling_unloading"),
        }

        row_num = header_row + 1

        lines = list(document.invoice_wh_lines.all())

        for line in lines:
            ws.cell(
                row=row_num,
                column=1,
                value=charge_type_map.get(line.charge_type, line.charge_type or ""),
            )
            ws.cell(row=row_num, column=2, value=_translated_line_description(line))
            ws.cell(row=row_num, column=3, value=float(line.quantity or 0))
            ws.cell(row=row_num, column=4, value=float(line.unit_price or 0))
            ws.cell(row=row_num, column=5, value=float(line.total or 0))
            row_num += 1

        total_row = row_num + 1

        ws.cell(row=total_row, column=4, value=_("total_amount")).font = Font(bold=True)
        ws.cell(row=total_row, column=5, value=float(document.total_amount or 0)).font = Font(bold=True)

        widths = {
            "A": 24,
            "B": 50,
            "C": 14,
            "D": 14,
            "E": 14,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        safe_name = (document.contact.company_name if document.contact else "billing_document").replace("/", "_")
        filename_dt = now.strftime("%Y%m%d_%H%M%S")

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="{safe_name}_{filename_dt}.xlsx"'
        )
        return response 


    @action(detail=False, methods=["get"], url_path="export-bills")
    def export_documents_invoices(self, request):
        company = get_user_company(request.user)

        qs = (
            WHBillingInvoice.objects
            .filter(company=company)
            .select_related("contact", "period")
            .prefetch_related("invoice_wh_lines", "invoice_wh_lines__charges")
            .order_by("-created_at", "-id")
        )

        period = request.query_params.get("period")
        contact = request.query_params.get("contact")
        status_param = request.query_params.get("status")
        period_from = request.query_params.get("period_from")
        period_to = request.query_params.get("period_to")



        if period:
            qs = qs.filter(period__uf=period)

        if contact:
            qs = qs.filter(contact__uf=contact)

        if status_param:
            qs = qs.filter(status=status_param)

        if period_from:
            qs = qs.filter(period__start_date__gte=period_from)

        if period_to:
            qs = qs.filter(period__end_date__lte=period_to) 

        wb = Workbook()
        ws = wb.active
        ws.title = _("billing_documents")

        user_tz = "Europe/Chisinau"
        timezone.activate(ZoneInfo(user_tz))
        now = timezone.localtime(timezone.now())

        ws["A1"] = _("billing_documents_export")
        ws["A1"].font = Font(bold=True, size=14)

        ws["A2"] = _("exported_at")
        ws["B2"] = now.strftime("%d-%m-%Y %H:%M")

        meta_row = 3

        if period_from:
            formatted_period_from = datetime.strptime(period_from, "%Y-%m-%d").strftime("%d-%m-%Y")
            ws.cell(row=meta_row, column=1, value=_("period_start"))
            ws.cell(row=meta_row, column=2, value=formatted_period_from)
            meta_row += 1

        if period_to:
            formatted_period_to = datetime.strptime(period_to, "%Y-%m-%d").strftime("%d-%m-%Y")
            ws.cell(row=meta_row, column=1, value=_("period_end"))
            ws.cell(row=meta_row, column=2, value=formatted_period_to)
            meta_row += 1

        header_row = meta_row + 2

        headers = [
            _("document_no"),
            _("customer"),
            _("period_start"),
            _("period_end"),
            _("status"),
            _("created_at"),
            _("total_amount"),
            _("line_type"),
            _("line_description"),
            _("quantity"),
            _("unit_price"),
            _("line_total"),
        ]

        header_row = 6
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = Font(bold=True)

        status_map = {
            "draft": _("draft"),
            "issued": _("issued"),
            "paid": _("paid"),
            "cancelled": _("cancelled"),
        }

        row_num = header_row + 1

        for document in qs:
            lines = list(document.invoice_wh_lines.all())

            if not lines:
                ws.cell(row=row_num, column=1, value=document.uf)
                ws.cell(
                    row=row_num,
                    column=2,
                    value=document.contact.company_name if document.contact else "",
                )
                ws.cell(
                    row=row_num,
                    column=3,
                    value=document.period.start_date.strftime("%d-%m-%Y") if document.period else "",
                )
                ws.cell(
                    row=row_num,
                    column=4,
                    value=document.period.end_date.strftime("%d-%m-%Y") if document.period else "",
                )
                ws.cell(
                    row=row_num,
                    column=5,
                    value=status_map.get(document.status, document.status or ""),
                )
                ws.cell(
                    row=row_num,
                    column=6,
                    value=timezone.localtime(document.created_at).strftime("%d-%m-%Y %H:%M")
                    if document.created_at else "",
                )
                ws.cell(row=row_num, column=7, value=float(document.total_amount or 0))
                row_num += 1

                ws.cell(row=row_num, column=11, value=_("Invoice total")).font = Font(bold=True)
                ws.cell(row=row_num, column=12, value=float(document.total_amount or 0)).font = Font(bold=True)
                row_num += 2
                continue

            for line in lines:
                ws.cell(row=row_num, column=1, value=document.uf)
                ws.cell(
                    row=row_num,
                    column=2,
                    value=document.contact.company_name if document.contact else "",
                )
                ws.cell(
                    row=row_num,
                    column=3,
                    value=document.period.start_date.strftime("%d-%m-%Y") if document.period else "",
                )
                ws.cell(
                    row=row_num,
                    column=4,
                    value=document.period.end_date.strftime("%d-%m-%Y") if document.period else "",
                )
                ws.cell(
                    row=row_num,
                    column=5,
                    value=status_map.get(document.status, document.status or ""),
                )
                ws.cell(
                    row=row_num,
                    column=6,
                    value=timezone.localtime(document.created_at).strftime("%d-%m-%Y %H:%M")
                    if document.created_at else "",
                )
                ws.cell(row=row_num, column=7, value=float(document.total_amount or 0))
                ws.cell(
                    row=row_num,
                    column=8,
                    value=_translated_charge_type_label(line.charge_type),
                )
                ws.cell(
                    row=row_num,
                    column=9,
                    value=_translated_line_description(line),
                )
                ws.cell(row=row_num, column=10, value=float(line.quantity or 0))
                ws.cell(row=row_num, column=11, value=float(line.unit_price or 0))
                ws.cell(row=row_num, column=12, value=float(line.total or 0))

                row_num += 1

            ws.cell(row=row_num, column=11, value=_("invoice_total")).font = Font(bold=True)
            ws.cell(row=row_num, column=12, value=float(document.total_amount or 0)).font = Font(bold=True)
            row_num += 2


        widths = {
            "A": 34,
            "B": 20,
            "C": 14,
            "D": 14,
            "E": 12,
            "F": 18,
            "G": 12,
            "H": 20,
            "I": 50,
            "J": 12,
            "K": 12,
            "L": 14,
        }

        for col, width in widths.items():
            ws.column_dimensions[col].width = width

        output = BytesIO()
        wb.save(output)
        output.seek(0)

        filename_dt = now.strftime("%Y%m%d_%H%M%S")

        response = HttpResponse(
            output.getvalue(),
            content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
        response["Content-Disposition"] = (
            f'attachment; filename="billing_documents_{filename_dt}.xlsx"'
        )
        return response

    @action(detail=True, methods=["get"], url_path="export-act")
    def export_act_excel(self, request, uf=None):
        document = self.get_object()

        if document.status != "issued":
            return Response(
                {"detail": _("The act can be generated only for issued invoices.")},
                status=status.HTTP_400_BAD_REQUEST,
            )

        lang = request.query_params.get("lang", "ro")
        if lang not in ["ro", "en", "ru"]:
            lang = "ro"

        user_tz = "Europe/Chisinau"
        timezone.activate(ZoneInfo(user_tz))
        now = timezone.localtime(timezone.now())

        with override(lang):
            wb = Workbook()
            ws = wb.active
            ws.title = _("Act")

            # -----------------------------
            # PAGE SETUP - A4 PORTRAIT
            # -----------------------------
            ws.page_setup.paperSize = ws.PAPERSIZE_A4
            ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT

            # Force Excel to scale the print area
            ws.sheet_properties.pageSetUpPr = PageSetupProperties(fitToPage=True)

            # If you want everything on ONE page:
            ws.page_setup.fitToWidth = 1
            ws.page_setup.fitToHeight = 1

            # Alternative:
            # fit width to 1 page, height can continue to next pages
            # ws.page_setup.fitToWidth = 1
            # ws.page_setup.fitToHeight = 0

            ws.page_margins = PageMargins(
                left=0.25, right=0.25, top=0.35, bottom=0.35, header=0.15, footer=0.15
            )

            ws.sheet_view.showGridLines = False
            ws.print_options.horizontalCentered = True

            # Print settings
            ws.print_options.horizontalCentered = True
            ws.print_title_rows = "$1:$12"

            # -----------------------------
            # STYLES
            # -----------------------------
            thin = Side(style="thin", color="000000")
            border = Border(left=thin, right=thin, top=thin, bottom=thin)

            bold_font = Font(bold=True, size=11)
            title_font = Font(bold=True, size=14)
            normal_font = Font(size=11)

            center = Alignment(horizontal="center", vertical="center", wrap_text=True)
            left = Alignment(horizontal="left", vertical="center", wrap_text=True)
            right = Alignment(horizontal="right", vertical="center", wrap_text=True)

            # -----------------------------
            # COLUMN WIDTHS
            # A:F
            # -----------------------------
            widths = {
                "A": 6,
                "B": 42,
                "C": 12,
                "D": 14,
                "E": 16,
                "F": 18,
            }
            for col, width in widths.items():
                ws.column_dimensions[col].width = width

            # -----------------------------
            # TRANSLATED LABELS
            # -----------------------------
            title = _("act_of_execution_of_services")
            provider_label = _("provider")
            client_label = _("client")
            document_no_label = _("document_no")
            period_label = _("period")
            issued_on_label = _("issued_on")
            no_label = _("no")
            description_label = _("description")
            quantity_label = _("quantity")
            unit_price_label = _("unit_price")
            total_label = _("total")
            grand_total_label = _("grand_total")
            provider_sign_label = _("provider_signature")
            client_sign_label = _("client_signature")

            # -----------------------------
            # HEADER
            # -----------------------------
            ws.merge_cells("A1:F1")
            ws["A1"] = title
            ws["A1"].font = title_font
            ws["A1"].alignment = center

            ws.merge_cells("A3:F3")
            ws["A3"] = f"{document_no_label}: {document.uf}"
            ws["A3"].font = bold_font
            ws["A3"].alignment = center

            ws["A5"] = provider_label
            ws["A5"].font = bold_font
            ws["B5"] = getattr(get_user_company(request.user), "name", "") or getattr(get_user_company(request.user), "title", "") or ""

            ws["A6"] = client_label
            ws["A6"].font = bold_font
            ws["B6"] = document.contact.company_name if document.contact else ""

            ws["D5"] = issued_on_label
            ws["D5"].font = bold_font
            ws["E5"] = now.strftime("%d-%m-%Y")

            ws["D6"] = period_label
            ws["D6"].font = bold_font
            ws["E6"] = (
                f"{document.period.start_date.strftime('%d-%m-%Y')} → "
                f"{document.period.end_date.strftime('%d-%m-%Y')}"
                if document.period else ""
            )

            # -----------------------------
            # TABLE HEADER
            # -----------------------------
            table_header_row = 9
            headers = [
                no_label,
                description_label,
                quantity_label,
                unit_price_label,
                total_label,
                "",
            ]

            for idx, header in enumerate(headers, start=1):
                cell = ws.cell(row=table_header_row, column=idx, value=header)
                cell.font = bold_font
                cell.alignment = center
                cell.border = border

            # Merge E/F visually for total column header
            ws.merge_cells(start_row=table_header_row, start_column=5, end_row=table_header_row, end_column=6)

            # -----------------------------
            # BODY
            # -----------------------------
            lines = list(document.invoice_wh_lines.all())
            row_num = table_header_row + 1

            for idx, line in enumerate(lines, start=1):
                description = _translated_line_description(line)

                ws.cell(row=row_num, column=1, value=idx)
                ws.cell(row=row_num, column=2, value=description)
                ws.cell(row=row_num, column=3, value=float(line.quantity or 0))
                ws.cell(row=row_num, column=4, value=float(line.unit_price or 0))
                ws.merge_cells(start_row=row_num, start_column=5, end_row=row_num, end_column=6)
                ws.cell(row=row_num, column=5, value=float(line.total or 0))

                for col_idx in range(1, 7):
                    cell = ws.cell(row=row_num, column=col_idx)
                    cell.font = normal_font
                    cell.border = border
                    cell.alignment = left if col_idx == 2 else center

                ws.cell(row=row_num, column=3).alignment = right
                ws.cell(row=row_num, column=4).alignment = right
                ws.cell(row=row_num, column=5).alignment = right

                ws.row_dimensions[row_num].height = 28
                row_num += 1

            # -----------------------------
            # TOTAL
            # -----------------------------
            total_row = row_num + 1

            ws.merge_cells(start_row=total_row, start_column=1, end_row=total_row, end_column=4)
            ws.cell(row=total_row, column=1, value=grand_total_label)
            ws.merge_cells(start_row=total_row, start_column=5, end_row=total_row, end_column=6)
            ws.cell(row=total_row, column=5, value=float(document.total_amount or 0))

            ws.cell(row=total_row, column=1).font = bold_font
            ws.cell(row=total_row, column=1).alignment = right
            ws.cell(row=total_row, column=5).font = bold_font
            ws.cell(row=total_row, column=5).alignment = right

            for col_idx in range(1, 7):
                ws.cell(row=total_row, column=col_idx).border = border

            # -----------------------------
            # SIGNATURES
            # -----------------------------
            sign_row = total_row + 4

            ws.merge_cells(start_row=sign_row, start_column=1, end_row=sign_row, end_column=2)
            ws.cell(row=sign_row, column=1, value=provider_label)
            ws.cell(row=sign_row, column=1).font = bold_font
            ws.cell(row=sign_row, column=1).alignment = center

            ws.merge_cells(start_row=sign_row, start_column=5, end_row=sign_row, end_column=6)
            ws.cell(row=sign_row, column=5, value=client_label)
            ws.cell(row=sign_row, column=5).font = bold_font
            ws.cell(row=sign_row, column=5).alignment = center

            line_row = sign_row + 3

            ws.merge_cells(start_row=line_row, start_column=1, end_row=line_row, end_column=2)
            ws.cell(row=line_row, column=1, value="________________________")
            ws.cell(row=line_row, column=1).alignment = center

            ws.merge_cells(start_row=line_row, start_column=5, end_row=line_row, end_column=6)
            ws.cell(row=line_row, column=5, value="________________________")
            ws.cell(row=line_row, column=5).alignment = center

            label_row = line_row + 1

            ws.merge_cells(start_row=label_row, start_column=1, end_row=label_row, end_column=2)
            ws.cell(row=label_row, column=1, value=provider_sign_label)
            ws.cell(row=label_row, column=1).alignment = center

            ws.merge_cells(start_row=label_row, start_column=5, end_row=label_row, end_column=6)
            ws.cell(row=label_row, column=5, value=client_sign_label)
            ws.cell(row=label_row, column=5).alignment = center

            # -----------------------------
            # VERTICAL ALIGNMENT
            # -----------------------------
            for r in range(1, label_row + 1):
                ws.row_dimensions[r].height = max(ws.row_dimensions[r].height or 15, 20)

            # -----------------------------
            # RESPONSE
            # -----------------------------
            output = BytesIO()
            wb.save(output)
            output.seek(0)

            customer_name = (document.contact.company_name if document.contact else "act").replace("/", "_")
            filename_dt = now.strftime("%Y%m%d_%H%M%S")
            filename = f"act_services_{customer_name}_{lang}_{filename_dt}.xlsx"

            response = HttpResponse(
                output.getvalue(),
                content_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
            response["Content-Disposition"] = f'attachment; filename="{filename}"'
            return response
        

class WHBillingInvoiceViewSet(viewsets.ReadOnlyModelViewSet):

    serializer_class = WHBillingInvoiceSerializer
    lookup_field = "uf"

    def get_queryset(self):

        company = get_user_company(self.request.user)

        queryset = (
            WHBillingInvoice.objects
            .filter(company=company)
            .select_related("contact", "period")
            .prefetch_related(
                Prefetch(
                    "invoice_wh_lines",
                    queryset=WHBillingInvoiceLine.objects.all()
                )
            )
            .order_by("-created_at")
        )

        # Filters
        period = self.request.query_params.get("period")
        contact = self.request.query_params.get("contact")
        period_from = self.request.query_params.get("period_from")
        period_to = self.request.query_params.get("period_to")


        if period:
            queryset = queryset.filter(period__uf=period)

        if contact:
            queryset = queryset.filter(contact__uf=contact)
      
        if period_from:
            queryset = queryset.filter(period__start_date__gte=period_from)

        if period_to:
            queryset = queryset.filter(period__end_date__lte=period_to)
             

        return queryset
    
    
    @action(detail=False, methods=["post"], url_path="create-for-contact")
    @transaction.atomic
    def create_for_contact(self, request):

        company = get_user_company(request.user)

        contact_uf = request.data.get("contact")
        contact = get_object_or_404(Contact, uf=contact_uf)

    
        period = _get_or_create_current_period(company)
      
        print('1144', period)

        if not period:
            return Response({"detail": "No active billing period"}, status=400)

        if period.is_closed:
            return Response({"detail": "Billing period closed"}, status=400)

        # Generate storage charges ONLY for this contact
        generate_storage_billing_for_period(
            company=company,
            period=period,
            contact_ids=[contact.id],
        )

        print('1148', )


        charges = WHBillingCharge.objects.filter(
            company=company,
            contact=contact,
            billing_period=period,
            invoiced=False
        )

        print('1152', charges)

        if not charges.exists():
            return Response({"detail": "No charges to invoice"}, status=400)

        total = (
            charges.aggregate(s=Sum("total"))["s"] or Decimal("0")
        ).quantize(Decimal("0.01"))

        invoice = WHBillingInvoice.objects.create(
            company=company,
            contact=contact,
            period=period,
            total_amount=total
        )

        lines = []

        for charge in charges:

            line = WHBillingInvoiceLine.objects.create(
                invoice=invoice,
                charge_type=charge.charge_type,
                description = f"({period.start_date} → {period.end_date})",
                quantity=charge.quantity,
                unit_price=charge.unit_price,
                total=charge.total
            )

            line.charges.add(charge)

            charge.invoiced = True
            charge.save(update_fields=["invoiced"])

            lines.append(line)

        return Response({
            "invoice": invoice.uf,
            "lines": len(lines),
            "total": str(total)
        })


    @action(detail=True, methods=["post"])
    def issue(self, request, uf=None):

        invoice = self.get_object()

        if invoice.status != "draft":
            return Response({"detail": "Only draft invoices can be issued"}, status=400)

        invoice.status = "issued"
        invoice.save(update_fields=["status"])

        return Response({"status": "issued"})
    
    # If view is ReadOnlyModelViewSet must add destroy:
    @transaction.atomic
    def destroy(self, request, *args, **kwargs):
        invoice = self.get_object()

        if invoice.status != "draft":
            return Response(
                {"detail": "Only draft invoices can be deleted"},
                status=status.HTTP_400_BAD_REQUEST
            )

        # unmark linked charges
        charge_ids = list(
            WHBillingCharge.objects.filter(invoice_lines__invoice=invoice).values_list("id", flat=True)
        )

        if charge_ids:
            WHBillingCharge.objects.filter(id__in=charge_ids).update(invoiced=False)

        invoice.delete()

        return Response(status=status.HTTP_204_NO_CONTENT)
    
    @action(detail=True, methods=["post"], url_path="manual-lines")
    @transaction.atomic
    def create_manual_lines(self, request, uf=None):
        invoice = self.get_object()

        if invoice.status != "draft":
            return Response(
                {"detail": "Manual lines can be added only to draft invoices"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        serializer = WHBillingInvoiceManualLinesCreateSerializer(data=request.data)
        serializer.is_valid(raise_exception=True)

        created_lines = []
        added_total = Decimal("0")

        for item in serializer.validated_data["lines"]:
            quantity = item["quantity"]
            unit_price = item["unit_price"]
            total = quantity * unit_price

            line = WHBillingInvoiceLine.objects.create(
                invoice=invoice,
                charge_type="manual",
                description=item["description"],
                quantity=quantity,
                unit_price=unit_price,
                total=total,
                is_manual=True,
            )
            created_lines.append(line.uf)
            added_total += total

        invoice.total_amount = (Decimal(invoice.total_amount or 0) + added_total).quantize(Decimal("0.01"))
        invoice.save(update_fields=["total_amount"])

        return Response(
            {
                "detail": "Manual lines created successfully",
                "lines": created_lines,
                "invoice_total_amount": str(invoice.total_amount),
            },
            status=status.HTTP_201_CREATED,
        )
    
    @action(detail=True, methods=["delete"], url_path=r"manual-lines/(?P<line_uf>[^/.]+)")
    @transaction.atomic
    def delete_manual_line(self, request, uf=None, line_uf=None):
        invoice = self.get_object()

        if invoice.status != "draft":
            return Response(
                {"detail": "Manual lines can be deleted only from draft invoices"},
                status=status.HTTP_400_BAD_REQUEST,
            )

        line = get_object_or_404(
            WHBillingInvoiceLine,
            invoice=invoice,
            uf=line_uf,
            is_manual=True,
        )

        line_total = Decimal(line.total or 0)
        line.delete()

        invoice.total_amount = max(
            Decimal("0.00"),
            (Decimal(invoice.total_amount or 0) - line_total).quantize(Decimal("0.01"))
        )
        invoice.save(update_fields=["total_amount"])

        return Response(
            {
                "detail": "Manual line deleted successfully",
                "invoice_total_amount": str(invoice.total_amount),
            },
            status=status.HTTP_200_OK,
        )


